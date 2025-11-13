"""Core macro checking functionality."""

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Tuple

from claude_agent_sdk import AssistantMessage, ClaudeAgentOptions, TextBlock, query
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

try:
    from oletools.olevba import VBA_Parser

    HAS_OLETOOLS = True
except ImportError:
    HAS_OLETOOLS = False

try:
    from odf.opendocument import load as load_odf
    from odf.table import Table, TableCell, TableRow
    from odf.text import P

    HAS_ODFPY = True
except ImportError:
    HAS_ODFPY = False


@dataclass
class MacroFinding:
    """Represents a discovered macro or suspicious code."""

    item_number: int
    location: str  # e.g., "Sheet1!A1" or "VBA Module: Module1"
    code: str
    score: int  # 1-10, where 1=malicious, 10=safe
    analysis: str
    cell_reference: Optional[Tuple[str, str, int]] = (
        None  # (sheet_name, column_letter, row)
    )


class MacroChecker:
    """Main class for checking macros in spreadsheet files."""

    def __init__(self, input_file: str, remove_threshold: int = 5):
        self.input_file = Path(input_file)
        self.remove_threshold = remove_threshold
        self.findings: List[MacroFinding] = []
        self.workbook = None
        self.ods_doc = None
        self.item_counter = 0
        self.file_type = None  # 'excel' or 'ods'

    def load_spreadsheet(self):
        """Load the spreadsheet (Excel or OpenOffice)."""
        try:
            suffix = self.input_file.suffix.lower()

            if suffix in [".xlsx", ".xlsm"]:
                self.file_type = "excel"
                self.workbook = load_workbook(self.input_file, data_only=False)
                print(f"Loaded Excel spreadsheet: {self.input_file}")
                return True
            elif suffix == ".ods":
                if not HAS_ODFPY:
                    print("Error: odfpy not installed. Install with: pip install odfpy")
                    return False
                self.file_type = "ods"
                self.ods_doc = load_odf(str(self.input_file))
                print(f"Loaded OpenOffice spreadsheet: {self.input_file}")
                return True
            else:
                print(f"Unsupported file format: {suffix}")
                return False

        except Exception as e:
            print(f"Error loading spreadsheet: {e}")
            return False

    def detect_vba_macros(self) -> List[Tuple[str, str]]:
        """Detect VBA macros in the spreadsheet."""
        macros = []

        if not HAS_OLETOOLS:
            print("Warning: oletools not installed, VBA macro detection disabled")
            print("Install with: pip install oletools")
            return macros

        try:
            vba_parser = VBA_Parser(str(self.input_file))

            if vba_parser.detect_vba_macros():
                print("VBA macros detected!")

                for (
                    _filename,
                    _stream_path,
                    vba_filename,
                    vba_code,
                ) in vba_parser.extract_macros():
                    if vba_code:
                        location = f"VBA Module: {vba_filename}"
                        macros.append((location, vba_code))

            vba_parser.close()
        except Exception as e:
            print(f"Error detecting VBA macros: {e}")

        return macros

    def detect_formula_cells(self) -> List[Tuple[str, str, str, int]]:
        """Detect cells with formulas (Excel and ODS)."""
        formula_cells = []

        if self.file_type == "excel":
            return self._detect_excel_formulas()
        elif self.file_type == "ods":
            return self._detect_ods_formulas()

        return formula_cells

    def _detect_excel_formulas(self) -> List[Tuple[str, str, str, int]]:
        """Detect formulas in Excel spreadsheets."""
        formula_cells = []

        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]

            for row in sheet.iter_rows():
                for cell in row:
                    if (
                        cell.value
                        and isinstance(cell.value, str)
                        and cell.value.startswith("=")
                    ):
                        formula = cell.value
                        location = f"{sheet_name}!{cell.coordinate}"
                        formula_cells.append(
                            (location, formula, sheet_name, cell.column, cell.row)
                        )

        return formula_cells

    def _detect_ods_formulas(self) -> List[Tuple[str, str, str, int]]:
        """Detect formulas in OpenOffice spreadsheets."""
        formula_cells = []

        tables = self.ods_doc.spreadsheet.getElementsByType(Table)

        for table in tables:
            sheet_name = table.getAttribute("name") or "Sheet"
            rows = table.getElementsByType(TableRow)

            row_num = 0
            for row in rows:
                row_num += 1
                cells = row.getElementsByType(TableCell)

                col_num = 0
                for cell in cells:
                    col_num += 1

                    # Handle repeated columns
                    repeat = cell.getAttribute("numbercolumnsrepeated")
                    if repeat:
                        try:
                            repeat_count = int(repeat)
                            # Skip large repeated empty columns
                            if repeat_count > 100:
                                col_num += repeat_count - 1
                                continue
                        except (ValueError, TypeError):
                            pass

                    # Get formula
                    formula = cell.getAttribute("formula")

                    if formula:
                        # Convert column number to letter
                        col_letter = get_column_letter(col_num)
                        location = f"{sheet_name}!{col_letter}{row_num}"
                        formula_cells.append(
                            (location, formula, sheet_name, col_num, row_num)
                        )

        return formula_cells

    async def analyze_code_with_claude(
        self, code: str, location: str
    ) -> Tuple[int, str]:
        """Use Claude SDK to analyze code and return a score and analysis.

        Returns:
            Tuple of (score: int 1-10, analysis: str)
        """
        prompt = f"""Analyze the following code/macro found in a spreadsheet for security risks.

Location: {location}

Code:
```
{code}
```

Provide:
1. A security score from 1-10 where:
   - 1-3: Definitely malicious (file access, network calls, process execution, obfuscation)
   - 4-6: Suspicious (external references, dynamic execution, questionable patterns)
   - 7-9: Potentially risky but may be legitimate (common functions that could be misused)
   - 10: Safe (simple calculations, harmless formulas)

2. A brief analysis explaining the score (2-3 sentences)

Format your response EXACTLY as:
SCORE: <number>
ANALYSIS: <your analysis here>
"""

        options = ClaudeAgentOptions(
            max_turns=1,
            system_prompt="You are a security analyst specializing in spreadsheet macro and formula analysis. Be concise and precise.",
        )

        score = 5  # default
        analysis = "Unable to analyze"

        try:
            async for message in query(prompt=prompt, options=options):
                if isinstance(message, AssistantMessage):
                    for block in message.content:
                        if isinstance(block, TextBlock):
                            text = block.text

                            # Parse score and analysis
                            lines = text.strip().split("\n")
                            for line in lines:
                                if line.startswith("SCORE:"):
                                    try:
                                        score = int(line.replace("SCORE:", "").strip())
                                        score = max(1, min(10, score))  # Clamp to 1-10
                                    except ValueError:
                                        pass
                                elif line.startswith("ANALYSIS:"):
                                    analysis = line.replace("ANALYSIS:", "").strip()

                            # If analysis wasn't on same line, get remaining text
                            if analysis == "Unable to analyze" and "ANALYSIS:" in text:
                                analysis = text.split("ANALYSIS:")[1].strip()

        except Exception as e:
            print(f"Error analyzing with Claude: {e}")
            analysis = f"Error during analysis: {str(e)}"

        return score, analysis

    async def scan_file(self):
        """Main scanning function."""
        if not self.load_spreadsheet():
            return False

        print("\n=== Scanning for macros and suspicious code ===\n")

        # 1. Detect VBA macros
        print("Checking for VBA macros...")
        vba_macros = self.detect_vba_macros()

        for location, code in vba_macros:
            self.item_counter += 1
            print(f"Analyzing VBA macro {self.item_counter}: {location}...")

            score, analysis = await self.analyze_code_with_claude(code, location)

            finding = MacroFinding(
                item_number=self.item_counter,
                location=location,
                code=code,
                score=score,
                analysis=analysis,
            )
            self.findings.append(finding)
            print(f"  Score: {score}/10")

        # 2. Detect suspicious formulas
        print("\nChecking for suspicious formulas...")
        formula_cells = self.detect_formula_cells()

        for location, formula, sheet_name, col, row in formula_cells:
            self.item_counter += 1
            print(f"Analyzing formula {self.item_counter}: {location}...")

            score, analysis = await self.analyze_code_with_claude(formula, location)

            col_letter = get_column_letter(col)
            finding = MacroFinding(
                item_number=self.item_counter,
                location=location,
                code=formula,
                score=score,
                analysis=analysis,
                cell_reference=(sheet_name, col_letter, row),
            )
            self.findings.append(finding)
            print(f"  Score: {score}/10")

        print(f"\n=== Scan complete: {len(self.findings)} items found ===\n")
        return True

    def generate_markdown_report(self) -> str:
        """Generate a markdown report of findings."""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        report = f"""# Macro Security Analysis Report

**File:** {self.input_file.name}
**Scan Date:** {timestamp}
**Total Items Found:** {len(self.findings)}
**Removal Threshold:** {self.remove_threshold}

## Summary

"""

        # Summary statistics
        malicious = len([f for f in self.findings if f.score <= 3])
        suspicious = len([f for f in self.findings if 4 <= f.score <= 6])
        risky = len([f for f in self.findings if 7 <= f.score <= 9])
        safe = len([f for f in self.findings if f.score == 10])

        report += f"- **Malicious (1-3):** {malicious}\n"
        report += f"- **Suspicious (4-6):** {suspicious}\n"
        report += f"- **Potentially Risky (7-9):** {risky}\n"
        report += f"- **Safe (10):** {safe}\n"
        report += f"- **Items to be removed (score < {self.remove_threshold}):** "
        report += (
            f"{len([f for f in self.findings if f.score < self.remove_threshold])}\n\n"
        )

        # Detailed findings
        report += "## Detailed Findings\n\n"

        for finding in sorted(self.findings, key=lambda x: x.score):
            report += f"### Item #{finding.item_number}: {finding.location}\n\n"
            report += f"**Score:** {finding.score}/10\n\n"
            report += f"**Analysis:** {finding.analysis}\n\n"
            report += f"**Code:**\n```\n{finding.code[:500]}"
            if len(finding.code) > 500:
                report += "\n... (truncated)"
            report += "\n```\n\n"
            report += "---\n\n"

        return report

    def create_sanitized_copy(self, output_file: Path):
        """Create a copy of the spreadsheet with suspicious cells highlighted/removed."""
        if self.file_type == "excel":
            return self._create_sanitized_excel(output_file)
        elif self.file_type == "ods":
            return self._create_sanitized_ods(output_file)
        return False

    def _create_sanitized_excel(self, output_file: Path):
        """Create sanitized copy of Excel file."""
        try:
            # Create a copy of the workbook
            output_wb = load_workbook(self.input_file)

            # Yellow fill for highlighted cells
            yellow_fill = PatternFill(
                start_color="FFFF00", end_color="FFFF00", fill_type="solid"
            )

            items_removed = 0

            for finding in self.findings:
                if finding.score < self.remove_threshold and finding.cell_reference:
                    sheet_name, col_letter, row = finding.cell_reference
                    sheet = output_wb[sheet_name]
                    cell = sheet[f"{col_letter}{row}"]

                    # Replace with reference and highlight
                    cell.value = f"CODE REMOVED: Item #{finding.item_number}"
                    cell.fill = yellow_fill
                    items_removed += 1

            # Save the sanitized copy
            output_wb.save(output_file)
            print(f"\nSanitized copy created: {output_file}")
            print(f"Items removed/highlighted: {items_removed}")

            return True

        except Exception as e:
            print(f"Error creating sanitized copy: {e}")
            return False

    def _create_sanitized_ods(self, output_file: Path):
        """Create sanitized copy of ODS file."""
        try:
            # Import needed for creating ODS elements
            import shutil

            from odf.style import Style, TableCellProperties

            # Copy the original file
            shutil.copy2(self.input_file, output_file)

            # Load the copy
            output_doc = load_odf(str(output_file))

            # Create yellow background style
            yellow_style = Style(name="YellowBackground", family="table-cell")
            yellow_style.addElement(TableCellProperties(backgroundcolor="#FFFF00"))
            output_doc.automaticstyles.addElement(yellow_style)

            items_removed = 0

            # Create a mapping of cells to remove
            cells_to_sanitize = {}
            for finding in self.findings:
                if finding.score < self.remove_threshold and finding.cell_reference:
                    sheet_name, col_letter, row = finding.cell_reference
                    key = (sheet_name, row)
                    if key not in cells_to_sanitize:
                        cells_to_sanitize[key] = []
                    cells_to_sanitize[key].append(
                        (finding.cell_reference[1], finding.item_number)
                    )

            # Process tables
            tables = output_doc.spreadsheet.getElementsByType(Table)
            for table in tables:
                sheet_name = table.getAttribute("name") or "Sheet"
                rows = table.getElementsByType(TableRow)

                row_num = 0
                for row in rows:
                    row_num += 1  # noqa: SIM113 - spreadsheet rows start at 1
                    cells = row.getElementsByType(TableCell)

                    col_num = 0
                    for cell in cells:
                        col_num += 1  # noqa: SIM113 - spreadsheet cols start at 1

                        # Check if this cell should be sanitized
                        key = (sheet_name, row_num)
                        if key in cells_to_sanitize:
                            col_letter = get_column_letter(col_num)
                            for target_col, item_num in cells_to_sanitize[key]:
                                if col_letter == target_col:
                                    # Remove formula
                                    cell.removeAttribute("formula")

                                    # Set text content
                                    # Remove existing text
                                    for p in cell.getElementsByType(P):
                                        cell.removeChild(p)

                                    # Add new text
                                    from odf.text import P as Paragraph

                                    p = Paragraph()
                                    p.addText(f"CODE REMOVED: Item #{item_num}")
                                    cell.addElement(p)

                                    # Apply yellow style
                                    cell.setAttribute("stylename", yellow_style)
                                    items_removed += 1

            # Save the sanitized copy
            output_doc.save(output_file)
            print(f"\nSanitized copy created: {output_file}")
            print(f"Items removed/highlighted: {items_removed}")

            return True

        except Exception as e:
            print(f"Error creating sanitized ODS copy: {e}")
            import traceback

            traceback.print_exc()
            return False
